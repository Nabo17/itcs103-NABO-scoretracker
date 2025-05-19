import openpyxl
from openpyxl import Workbook
import os

LEDGER_FILE = "student_scores.xlsx"

def initialize_scorebook():
    if not os.path.exists(LEDGER_FILE):
        new_workbook = Workbook()
        sheet = new_workbook.active
        sheet.title = "Scores"
        sheet.append(["Student Name", "Score", "Status"])
        new_workbook.save(LEDGER_FILE)
        print("📗 New scorebook initialized!\n")

def log_student_performance(student_name, score):
    try:
        numeric_score = float(score)
    except ValueError:
        print("⚠️ Invalid score format. Please enter a number.")
        return

    status = "Pass" if numeric_score >= 50 else "Fail"

    workbook = openpyxl.load_workbook(LEDGER_FILE)
    worksheet = workbook["Scores"]

    updated = False
    for row in worksheet.iter_rows(min_row=2):
        if row[0].value == student_name:
            row[1].value = numeric_score
            row[2].value = status
            updated = True
            print(f"🔄 {student_name}'s record updated to {numeric_score} - {status}")
            break

    if not updated:
        worksheet.append([student_name, numeric_score, status])
        print(f"➕ Added new student: {student_name} - {numeric_score} ({status})")

    workbook.save(LEDGER_FILE)

def display_all_records():
    if not os.path.exists(LEDGER_FILE):
        print("🚫 No scorebook found. Please add a record first.")
        return

    workbook = openpyxl.load_workbook(LEDGER_FILE)
    worksheet = workbook["Scores"]

    print("\n🎓 Academic Leaderboard")
    print(f"{'Name':<20}{'Score':<10}{'Status':<10}")
    print("-" * 40)
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        print(f"{row[0]:<20}{row[1]:<10}{row[2]:<10}")
    print()

def command_center():
    initialize_scorebook()

    while True:
        print("\n=== Student Score Management ===")
        print("1. Add or Update Score")
        print("2. View All Records")
        print("3. Exit Program")
        choice = input("Select an option (1/2/3): ").strip()

        if choice == "1":
            name = input("Enter student's name: ").strip()
            score = input("Enter student's score (0-100): ").strip()
            try:
                score_val = float(score)
                if 0 <= score_val <= 100:
                    log_student_performance(name, score_val)
                else:
                    print("❗ Score must be between 0 and 100.")
            except ValueError:
                print("❗ Please enter a valid numeric score.")
        elif choice == "2":
            display_all_records()
        elif choice == "3":
            print("👋 Exiting program. Have a productive day!")
            break
        else:
            print("❌ Invalid choice. Please try again.")

if __name__ == "__main__":
    command_center()
